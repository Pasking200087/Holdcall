"""
excel.py — Импорт и экспорт контактов через Excel (openpyxl)

Формат импорта (первая строка — заголовки):
  Обязательно: Телефон (или Phone)
  Опционально:  Имя / Name, Компания / Company, Комментарий / Comment
"""
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import STATUS_LABELS


# Варианты названий колонок (регистронезависимо)
_COL_ALIASES = {
    "phone":   ["телефон", "phone", "тел", "номер", "tel", "mobile"],
    "name":    ["имя", "name", "фио", "контакт", "contact"],
    "company": ["компания", "company", "организация", "org", "firm"],
    "comment": ["комментарий", "comment", "примечание", "note", "заметка"],
}


def _normalize_phone(raw: str) -> str:
    """Убрать лишние символы из номера."""
    digits = re.sub(r"[^\d+]", "", str(raw).strip())
    return digits if digits else str(raw).strip()


def _detect_columns(headers: list[str]) -> dict[str, int]:
    """Сопоставить заголовки с полями. Возвращает {field: col_index}."""
    mapping: dict[str, int] = {}
    for idx, h in enumerate(headers):
        h_low = h.strip().lower()
        for field, aliases in _COL_ALIASES.items():
            if h_low in aliases and field not in mapping:
                mapping[field] = idx
    return mapping


def import_from_excel(path: str) -> tuple[list[dict], list[str]]:
    """
    Прочитать Excel и вернуть (rows, errors).
    rows — список dict {name, phone, company, comment}
    errors — список строк с описанием пропущенных строк
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # Первая строка — заголовки
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Файл пуст"]

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    mapping = _detect_columns(headers)

    if "phone" not in mapping:
        # Попробовать без заголовков — считать первую колонку телефоном
        # (если в первой строке уже номер)
        first_val = headers[0] if headers else ""
        if re.search(r"\d{7,}", first_val.replace(" ", "")):
            # Заголовков нет, первая строка — данные
            mapping = {"phone": 0}
            # Вернуть итератор обратно, включив header_row
            all_rows = [header_row] + list(rows_iter)
        else:
            return [], [
                "Не найдена колонка 'Телефон'. "
                "Убедитесь, что в первой строке есть заголовок 'Телефон'."
            ]
    else:
        all_rows = list(rows_iter)

    result: list[dict] = []
    errors: list[str] = []

    for row_num, row in enumerate(all_rows, start=2):
        try:
            phone_raw = row[mapping["phone"]] if mapping["phone"] < len(row) else None
            if phone_raw is None or str(phone_raw).strip() == "":
                errors.append(f"Строка {row_num}: пустой телефон, пропущено")
                continue

            phone = _normalize_phone(str(phone_raw))
            name    = str(row[mapping["name"]]).strip()    if "name"    in mapping and row[mapping["name"]]    is not None else ""
            company = str(row[mapping["company"]]).strip() if "company" in mapping and row[mapping["company"]] is not None else ""
            comment = str(row[mapping["comment"]]).strip() if "comment" in mapping and row[mapping["comment"]] is not None else ""

            result.append({
                "phone":   phone,
                "name":    name,
                "company": company,
                "comment": comment,
            })
        except Exception as e:
            errors.append(f"Строка {row_num}: {e}")

    wb.close()
    return result, errors


def export_to_excel(contacts: list[dict], path: str) -> None:
    """Сохранить список контактов в Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"

    # Стиль заголовков
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Имя", "Телефон", "Компания", "Статус",
               "Результат звонка", "Кто звонил", "Дата звонка", "Комментарий"]
    col_widths = [5, 25, 18, 22, 14, 35, 18, 16, 30]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20

    for row_idx, c in enumerate(contacts, start=2):
        status_label = STATUS_LABELS.get(c.get("status", ""), c.get("status", ""))
        caller = c.get("caller_name") or c.get("caller_username") or ""
        ws.append([
            row_idx - 1,
            c.get("name", ""),
            c.get("phone", ""),
            c.get("company", ""),
            status_label,
            c.get("call_result", ""),
            caller,
            c.get("call_date", ""),
            c.get("comment", ""),
        ])
        ws.row_dimensions[row_idx].height = 16

    # Закрепить первую строку
    ws.freeze_panes = "A2"

    wb.save(path)
